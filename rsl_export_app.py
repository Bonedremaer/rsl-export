# -*- coding: utf-8 -*-
"""RSL Export — Web-App (Streamlit Community Cloud / iPhone) — APP 1 von 2.

Nachbau der lokalen rsl_app.py (B5-Buffer-Ansicht), vollständig eigenständig
(keine lokalen Pfade), cloud-deploybar. Lädt live S&P-500-Kurse via yfinance,
RSL (Tagesschluss, SMA 130/200 HT) wie RSL_Update.py, Schock-Filter +
Regime-Ampel + Ziel-Depot mit Sektor-Cap, Top-20-Ranking und Excel-Download.
"""
import datetime, io, warnings
warnings.filterwarnings("ignore")
import pandas as pd
import numpy as np
import streamlit as st
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCHOCK_PCT = 0.04
SCHOCK_TAGE = 5
SEKTOR_CAP = 2
JUMP_MAX = 0.40

st.set_page_config(page_title="RSL Export", page_icon="📈", layout="centered")


@st.cache_data(show_spinner=False)
def lade_universum() -> pd.DataFrame:
    return pd.read_csv("sp500_universe.csv")


@st.cache_data(ttl=60 * 30, show_spinner=False)
def lade_alles():
    uni = lade_universum()
    sek = {r.Symbol: (r.Sektor, r.Untersektor) for r in uni.itertuples()}
    symbole = [s.replace(".", "-") for s in uni["Symbol"].tolist()]
    try:
        fx = float(yf.download("EURUSD=X", period="5d", progress=False,
                               auto_adjust=True)["Close"].dropna().iloc[-1])
    except Exception:
        fx = 1.17
    closes = {}
    BATCH = 120
    for i in range(0, len(symbole), BATCH):
        part = symbole[i:i + BATCH]
        raw = yf.download(part, period="350d", interval="1d",
                          auto_adjust=True, progress=False, threads=True)
        cl = raw["Close"]
        if isinstance(cl, pd.Series):
            cl = cl.to_frame(part[0])
        for s in cl.columns:
            closes[s] = cl[s]
    px = pd.DataFrame(closes).sort_index()
    rows, artefakte = [], []
    for s in px.columns:
        ser = px[s].dropna()
        if len(ser) < 130:
            continue
        if float(ser.iloc[-20:].pct_change().abs().max()) > JUMP_MAX:
            artefakte.append(s.replace("-", "."))
            continue
        price = ser.iloc[-1]
        sma40 = ser.rolling(200, min_periods=200).mean().iloc[-1]
        sma26 = ser.rolling(130, min_periods=130).mean().iloc[-1]
        if not np.isfinite(sma26):
            continue
        orig = s.replace("-", ".")
        sk, sub = sek.get(s, sek.get(orig, ("", "")))
        rows.append(dict(
            Symbol=orig, Sektor=sk, Branche=sub,
            Preis_USD=round(float(price), 2),
            Preis_EUR=round(float(price) / fx, 2),
            RSL_26W=round(float(price / sma26), 4),
            RSL_40W=round(float(price / sma40), 4) if np.isfinite(sma40) else np.nan,
        ))
    df = pd.DataFrame(rows)
    df = (df[(df["RSL_26W"] > 0.01) & (df["RSL_26W"] < 5.0)]
          .sort_values("RSL_26W", ascending=False).reset_index(drop=True))
    df.insert(0, "Rang", df.index + 1)
    try:
        close = yf.download("^GSPC", period="350d", auto_adjust=True,
                            progress=False)["Close"].dropna()
        kurs = float(close.iloc[-1]); kurs5 = float(close.iloc[-SCHOCK_TAGE - 1])
        r5t = (kurs - kurs5) / kurs5
        sma100 = float(close.tail(100).mean()); sma200 = float(close.tail(200).mean())
        regime = "Bullish" if kurs > sma100 else ("Neutral" if kurs > sma200 else "Bearish")
        schock = {"aktiv": r5t < -SCHOCK_PCT, "r5t_pct": round(r5t * 100, 2),
                  "kurs": round(kurs, 2), "regime": regime}
    except Exception:
        schock = {"aktiv": False, "r5t_pct": None, "kurs": 0, "regime": "unbekannt"}
    return df, fx, artefakte, schock


def ziel_depot_mit_cap(df_sortiert, n=3, cap=SEKTOR_CAP):
    ziel, skips, zahl = [], [], {}
    for _, row in df_sortiert.iterrows():
        if len(ziel) >= n:
            break
        t, s = str(row["Symbol"]), str(row.get("Sektor", "?") or "?")
        if s != "?" and zahl.get(s, 0) >= cap:
            skips.append(t); continue
        ziel.append(t)
        if s != "?":
            zahl[s] = zahl.get(s, 0) + 1
    return ziel, skips


def regime_ampel(df):
    r = df["RSL_26W"].dropna(); r = r[(r > 0.01) & (r < 5.0)]
    if len(r) < 50:
        return None
    disp = float(r.nlargest(5).mean() - r.median())
    if disp >= 0.45:   status, sym, hin = "GRÜN", "🟢", "wöchentliche Prüfung erlaubt"
    elif disp < 0.35:  status, sym, hin = "ROT", "🔴", "beim Quartalsrhythmus bleiben"
    else:              status, sym, hin = "GELB", "🟡", "Quartalsrhythmus, nicht hochschalten"
    return {"disp": round(disp, 3), "status": status, "symbol": sym, "hinweis": hin}


def baue_excel(df, fx):
    heute = datetime.date.today().strftime("%Y-%m-%d")
    HEAD = PatternFill("solid", fgColor="1F2A37"); HF = Font(bold=True, color="FFFFFF", size=11)
    GR = Font(color="1A7F37", bold=True); RD = Font(color="B42318")
    THIN = Side(style="thin", color="D0D7DE")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN); CEN = Alignment(horizontal="center")

    def sheet(ws, data, title):
        ws.sheet_view.showGridLines = False
        ws["A1"] = title; ws["A1"].font = Font(bold=True, size=13, color="1F2A37")
        ws["A2"] = (f"Datenstand {heute} · RSL_26W = Kurs / SMA(130 HT) · EUR/USD {fx:.4f} · "
                    f"yfinance · Artefakt-Filter >{int(JUMP_MAX*100)}%/Tag")
        ws["A2"].font = Font(size=9, color="6E7781")
        for j, h in enumerate(["Rang", "Ticker", "Branche", "Sektor", "RSL 26W", "Kurs USD", "Kurs EUR"], 1):
            c = ws.cell(4, j, h); c.fill = HEAD; c.font = HF; c.alignment = CEN; c.border = BORDER
        for k, row in enumerate(data.itertuples(), start=5):
            for j, v in enumerate([row.Rang, row.Symbol, row.Branche, row.Sektor,
                                   row.RSL_26W, row.Preis_USD, row.Preis_EUR], 1):
                c = ws.cell(k, j, v); c.border = BORDER
                if j in (1, 5): c.alignment = CEN
                if j == 5:
                    c.font = GR if (isinstance(v, (int, float)) and v > 1) else RD
                    c.number_format = "0.000"
                if j in (6, 7): c.number_format = '#,##0.00'
        for j, w in enumerate([6, 9, 30, 22, 10, 12, 12], 1):
            ws.column_dimensions[chr(64 + j)].width = w
        ws.freeze_panes = "A5"

    wb = Workbook(); ws1 = wb.active; ws1.title = "Top 20"
    sheet(ws1, df.head(20), "RSL Top 20 — B5-Buffer (SMA 26W)")
    sheet(wb.create_sheet("Gesamtranking"), df, "RSL Gesamtranking S&P 500 (SMA 26W)")
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ══ UI ════════════════════════════════════════════════════════════════════════
st.markdown("# 📈 RSL Export")
st.caption(f"Stand: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')} Uhr")

if st.button("🔄 Daten jetzt aktualisieren", use_container_width=True, type="primary"):
    lade_alles.clear()

with st.spinner("Lade S&P-500-Daten von Yahoo Finance … (kann 1–2 Minuten dauern)"):
    df, fx, artefakte, schock = lade_alles()

if schock.get("aktiv"):
    st.error(f"🚨 **SCHOCK-FILTER AKTIV** — S&P 500 5T-Rendite: {schock['r5t_pct']} %  |  "
             "B5-Buffer: ALLE Positionen sofort verkaufen!")
else:
    farbe = {"Bullish": "🟢", "Neutral": "🟡", "Bearish": "🔴"}.get(schock.get("regime", ""), "⚪")
    st.info(f"✅ Schock-Filter inaktiv  |  S&P 500: ${schock.get('kurs', 0):,.2f}  |  "
            f"{farbe} Regime: **{schock.get('regime', '?')}**  |  5T-Rendite: {schock.get('r5t_pct', '?')} %")

st.success(f"✅ {len(df)} Aktien mit gültigem RSL geladen  ·  Stand {datetime.date.today().strftime('%d.%m.%Y')}")
st.write("")

st.markdown(f"### 🎯 B5-Buffer (26W · Top 3 · halten solange Rang ≤ 5 · max. {SEKTOR_CAP}/Sektor)")
ampel = regime_ampel(df)
if ampel:
    st.markdown(f"**{ampel['symbol']} Regime-Ampel: {ampel['status']}** · "
                f"Dispersion {ampel['disp']} (🟢 ≥ 0,45 · 🔴 < 0,35) — {ampel['hinweis']}")

colb1, colb2 = st.columns(2)
with colb1:
    exb = df[(df["RSL_26W"] > 0.01) & (df["RSL_26W"] < 5.0)].nlargest(60, "RSL_26W").reset_index(drop=True)
    kauf, cap_skips = ziel_depot_mit_cap(exb, n=3)
    puffer = [t for t in exb.iloc[3:5]["Symbol"].tolist() if t not in kauf]
    st.markdown(f"**🟢 Ziel-Depot (max. {SEKTOR_CAP}/Sektor):** `{'  ·  '.join(kauf)}`")
    if puffer:
        st.markdown(f"**🟡 Halte-Puffer (Rang 4–5):** `{'  ·  '.join(puffer)}`")
    if cap_skips:
        st.markdown(f"**⛔ Sektor-Cap übersprungen:** `{'  ·  '.join(cap_skips[:5])}`")
    st.caption("Halten solange Rang ≤ 5, sonst verkaufen und durch den bestplatzierten "
               f"freien Titel ersetzen (max. {SEKTOR_CAP} je Sektor). Prüfung quartalsweise "
               "(letzter Handels-Freitag, 20:00 Uhr); bei 🟢 Ampel optional wöchentlich.")
with colb2:
    st.download_button("📊 Top 20 + Ranking als Excel", baue_excel(df, fx),
                       file_name=f"RSL_Top20_{datetime.date.today()}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)
    st.download_button("⬇️ Gesamtranking als CSV",
                       df.drop(columns=["Preis_EUR"]).to_csv(index=False).encode("utf-8"),
                       file_name=f"rsl_export_{datetime.date.today()}.csv",
                       mime="text/csv", use_container_width=True)

st.markdown("### 📊 Top 20 nach RSL (26W)")
top = df.head(20)[["Rang", "Symbol", "Sektor", "RSL_26W", "Preis_USD", "Preis_EUR"]].copy()
top.columns = ["Rang", "Ticker", "Sektor", "RSL 26W", "Kurs USD", "Kurs EUR"]
st.dataframe(top, hide_index=True, use_container_width=True,
             column_config={
                 "RSL 26W": st.column_config.NumberColumn(format="%.3f"),
                 "Kurs USD": st.column_config.NumberColumn(format="%.2f"),
                 "Kurs EUR": st.column_config.NumberColumn(format="%.2f"),
             })

with st.expander(f"📁 Gesamtranking ({len(df)} Titel) anzeigen"):
    st.dataframe(df[["Rang", "Symbol", "Sektor", "Branche", "RSL_26W", "Preis_USD", "Preis_EUR"]],
                 hide_index=True, use_container_width=True)

if artefakte:
    st.caption("Aussortierte Kursartefakte (Split/Spin-off-Sprung > 40 %/Tag): " + ", ".join(artefakte))
st.caption(f"RSL_26W = Kurs / SMA(130 Handelstage) · EUR/USD {fx:.4f} · Quelle yfinance · "
           "Stichtagsbetrachtung ohne Depotkenntnis · keine Anlageberatung.")
